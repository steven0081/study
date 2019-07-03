import docx , os
import openpyxl
from docx import Document

file_name =  os.getcwd()+'\\' + '附件12.2019年省级工业互联网发展专项资金申报指南.docx'
xlsx_file = os.getcwd()+'\\' + 'test.xlsx'
wb = openpyxl.load_workbook(xlsx_file)
sheet = wb.get_sheet_by_name('Sheet1')
sheet.cell(row=1, column=1).value = 'gggggggggs!'
document = Document(file_name)
tables = document.tables
print(len(tables))
table = tables[0]
doc_text_new, doc_text_old = '',''
for i in range(0,len(table.rows)):
    row_cells = table.rows[i].cells
    for j in range(0, len(row_cells)):
        doc_text_new = table.cell(i, j).text
        if doc_text_new != doc_text_old:
            sheet.cell(row=i+1, column=j+1).value = table.cell(i, j).text
            doc_text_old = table.cell(i, j).text
        # print(table.cell(i,j).text)



wb.save(xlsx_file)
