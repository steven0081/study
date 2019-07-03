from selenium import webdriver
import openpyxl
import time ,os
from lxml import etree

#初始化excel文件
xlsx_file = os.getcwd()+'\\' + 'test.xlsx'
wb = openpyxl.load_workbook(xlsx_file)
sheet = wb.get_sheet_by_name('Sheet1')
#启动浏览器
browser = webdriver.Chrome()
browser.get('http://61.166.187.188/')
time.sleep(1)
html = browser.page_source
#print(html)
html = etree.HTML(html)
user_elem = browser.find_element_by_id('username')
user_elem.send_keys('xxhhxxcyc')
pw_elem = browser.find_element_by_id('password')
pw_elem.send_keys('123456')
time.sleep(1)
#点击登录
browser.find_element_by_xpath('//*[@id="app"]/div[2]/div[2]/div/div/div[1]/div[2]/div/div[1]/div/div[2]/button[2]/div').click()
time.sleep(1)
#点击项目管理
browser.find_element_by_xpath('//*[@id="app"]/div[2]/div[2]/main/div/div/div/div[2]/div/div/div/div[1]/div/div[3]/a').click()
time.sleep(1)
#点击项目详情
#item_NO = [177, 174, 171,173,173,175,173]
row_number = '1' #行号信息
browser.find_element_by_xpath('//*[@id="proj-index-card"]/div[2]/div[1]/table/tbody/tr['+row_number+']/td[10]/span[1]').click()
time.sleep(1)
item_number ='177' #项目编号
row, col = 1, 1 # 设置excel 文件中行号
print('----单位基本信息-----')
for i in range(2, 12):
    item_value = browser.find_element_by_xpath('//*[@id="app"]/div['+item_number+']/div/div/div[2]/form/div/div[2]/div['+str(i)+']/div/div/div[1]/div/input')\
    .get_attribute('value')
    sheet.cell(row=row, column=col).value = item_value
    col+=1
    print(item_value)

for j in range(16, 22):
    item_value = browser.find_element_by_xpath('//*[@id="app"]/div[' + item_number + ']/div/div/div[2]/form/div/div[2]/div[' + str(j) + ']/div/div/div[1]/div/input') \
        .get_attribute('value')
    sheet.cell(row=row, column=col).value = item_value
    col+= 1
    print(item_value)

print('-----项目基本信息-----')
for j in range(25, 29):
    item_value = browser.find_element_by_xpath('//*[@id="app"]/div[' + item_number + ']/div/div/div[2]/form/div/div[2]/div[' + str(j) + ']/div/div/div[1]/div/input') \
        .get_attribute('value')
    sheet.cell(row=row, column=col).value = item_value
    col+= 1
    print(item_value)

for j in range(32, 41):
    item_value = browser.find_element_by_xpath('//*[@id="app"]/div[' + item_number + ']/div/div/div[2]/form/div/div[2]/div[' + str(j) + ']/div/div/div[1]/div/input') \
        .get_attribute('value')
    sheet.cell(row=row, column=col).value = item_value
    col += 1
    print(item_value)
#取平台简介
item_value = browser.find_element_by_xpath('//*[@id="app"]/div[' + item_number + ']/div/div/div[2]/form/div/div[2]/div[41]/div/div/div[1]/div/textarea') \
        .get_attribute('value')
sheet.cell(row=row, column=col).value = item_value
col+=1
print(item_value)
#投资计划
item_value = browser.find_element_by_xpath('//*[@id="app"]/div[' + item_number + ']/div/div/div[2]/form/div/div[2]/div[42]/div/div/div[1]/div/textarea') \
        .get_attribute('value')
sheet.cell(row=row, column=col).value = item_value
col+=1
print(item_value)
for j in range(43, 45):
    item_value = browser.find_element_by_xpath('//*[@id="app"]/div[' + item_number + ']/div/div/div[2]/form/div/div[2]/div[' + str(j) + ']/div/div/div[1]/div/input') \
        .get_attribute('value')
    sheet.cell(row=row, column=col).value = item_value
    col += 1
    print(item_value)

wb.save(xlsx_file)

'''
comp_value = browser.find_element_by_xpath('//*[@id="app"]/div['+item_number+']/div/div/div[2]/form/div/div[2]/div[2]/div/div/div[1]/div/input')\
    .get_attribute('value')
print(comp_value)
#取统一社会信用代码
lxr_value = browser.find_element_by_xpath('//*[@id="app"]/div['+item_number+']/div/div/div[2]/form/div/div[2]/div[3]/div/div/div[1]/div/input')\
    .get_attribute('value')
print(lxr_value)
#取注册地址
address_value = browser.find_element_by_xpath('//*[@id="app"]/div['+item_number+']/div/div/div[2]/form/div/div[2]/div[4]/div/div/div[1]/div/input')\
    .get_attribute('value')
print(address_value)
#取注册时间
time_value = browser.find_element_by_xpath('//*[@id="app"]/div['+item_number+']/div/div/div[2]/form/div/div[2]/div[5]/div/div/div[1]/div/input')\
    .get_attribute('value')
print(time_value)
#取注册资金
money_value = browser.find_element_by_xpath('//*[@id="app"]/div['+item_number+']/div/div/div[2]/form/div/div[2]/div[6]/div/div/div[1]/div/input')\
    .get_attribute('value')
print(money_value)
#取联系人
'''




